import os

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.utils import setup_logger

logger = setup_logger()


def generate_report(df, output_file):

    output_dir = os.path.dirname(output_file)

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    df.to_excel(
        output_file,
        sheet_name="Processed Data",
        index=False,
    )

    workbook = load_workbook(output_file)

    worksheet = workbook["Processed Data"]

    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    kpi_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    title_font = Font(
        bold=True,
        size=12,
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ==========================
    # Processed Data
    # ==========================

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column_cells in worksheet.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column_cells
        )

        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 3

    table = Table(
        displayName="SalesTable",
        ref=worksheet.dimensions,
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
    )

    table.tableStyleInfo = style

    worksheet.add_table(table)

    # ==========================
    # Summary
    # ==========================

    summary = workbook.create_sheet("Executive Dashboard")

    summary.freeze_panes = "A2"

    summary["A1"] = "Metric"
    summary["B1"] = "Value"

    metrics = [
        ("Transactions", len(df)),
        ("Total Sales", df["Total"].sum()),
        ("Average Sale", df["Total"].mean()),
        ("Highest Sale", df["Total"].max()),
        ("Lowest Sale", df["Total"].min()),
    ]

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, (metric, value) in enumerate(metrics, start=2):

        summary[f"A{row}"] = metric
        summary[f"B{row}"] = round(value, 2)

        summary[f"A{row}"].font = title_font
        summary[f"A{row}"].fill = kpi_fill

        summary[f"A{row}"].border = thin_border
        summary[f"B{row}"].border = thin_border

        summary[f"A{row}"].alignment = Alignment(horizontal="left")
        summary[f"B{row}"].alignment = Alignment(horizontal="right")

    summary["B3"].number_format = "$#,##0.00"
    summary["B4"].number_format = "$#,##0.00"
    summary["B5"].number_format = "$#,##0.00"
    summary["B6"].number_format = "$#,##0.00"

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18

    # ==========================
    # Sales by Product
    # ==========================

    sales = (
        df.groupby("Product")["Total"]
        .sum()
        .reset_index()
    )

    summary["D1"] = "Product"
    summary["E1"] = "Sales"

    for cell in summary["D1:E1"][0]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for index, row in sales.iterrows():

        summary[f"D{index + 2}"] = row["Product"]
        summary[f"E{index + 2}"] = row["Total"]

        summary[f"D{index + 2}"].border = thin_border
        summary[f"E{index + 2}"].border = thin_border

        summary[f"E{index + 2}"].number_format = "$#,##0.00"

    summary.column_dimensions["D"].width = 22
    summary.column_dimensions["E"].width = 18

    data = Reference(
        summary,
        min_col=5,
        min_row=1,
        max_row=len(sales) + 1,
    )

    categories = Reference(
        summary,
        min_col=4,
        min_row=2,
        max_row=len(sales) + 1,
    )

    chart = BarChart()

    chart.title = "Sales by Product"

    chart.y_axis.title = "Sales ($)"

    chart.x_axis.title = "Product"

    chart.style = 10

    chart.height = 10

    chart.width = 18

    chart.add_data(
        data,
        titles_from_data=True,
    )

    chart.set_categories(categories)

    summary.add_chart(
        chart,
        "G2",
    )

    workbook.save(output_file)

    logger.info(
        f"Report generated: {output_file}"
    )

    print("\nReport generated successfully.")
    print(f"Output file: {output_file}")